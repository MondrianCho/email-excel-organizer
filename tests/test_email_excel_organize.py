import sys
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from email_excel_organize import (
    COL_E,
    COL_I,
    COL_K,
    COL_L,
    COL_M,
    get_new_filename,
    is_email_file,
    is_excel_file,
    match_e_rules,
    normalize_title,
    parse_date_from_text,
    process_sheet,
    safe_str,
)


class TestSafeStr(unittest.TestCase):
    def test_none_returns_empty_string(self):
        self.assertEqual(safe_str(None), "")

    def test_strips_whitespace(self):
        self.assertEqual(safe_str("  hello  "), "hello")

    def test_non_string_is_stringified(self):
        self.assertEqual(safe_str(123), "123")


class TestIsExcelFile(unittest.TestCase):
    def test_accepts_xlsx(self):
        self.assertTrue(is_excel_file(Path("report.xlsx")))

    def test_accepts_xlsm(self):
        self.assertTrue(is_excel_file(Path("report.xlsm")))

    def test_rejects_other_extensions(self):
        self.assertFalse(is_excel_file(Path("report.csv")))

    def test_rejects_temp_lock_files(self):
        self.assertFalse(is_excel_file(Path("~$report.xlsx")))


class TestMatchERules(unittest.TestCase):
    def test_ack_keyword_is_case_insensitive(self):
        self.assertEqual(match_e_rules("automatic reply: out of office"), ("ACK", None, None, None))

    def test_maintenance_reminder_takes_priority_over_reminder(self):
        k, l, n, o = match_e_rules("Maintenance Reminder for patent 12345")
        self.assertEqual((k, l), ("비용", "(비용)연차료/RMD"))

    def test_annuity_keyword_maps_to_cost_category(self):
        k, l, n, o = match_e_rules("Annuity payment due")
        self.assertEqual(k, "비용")

    def test_no_match_returns_none(self):
        self.assertIsNone(match_e_rules("전혀 관련 없는 제목"))


class TestGetNewFilename(unittest.TestCase):
    def test_uses_current_date_when_not_using_mod_date(self):
        from datetime import datetime

        name = get_new_filename(Path("dummy.xlsx"), use_mod_date=False)
        today = datetime.now().strftime("%Y-%m-%d")
        self.assertEqual(name, f"{today} 이메일 접수_초안.xlsx")


class TestIsEmailFile(unittest.TestCase):
    def test_accepts_eml_and_msg(self):
        self.assertTrue(is_email_file(Path("a.eml")))
        self.assertTrue(is_email_file(Path("a.msg")))

    def test_rejects_other_extensions(self):
        self.assertFalse(is_email_file(Path("a.txt")))


class TestNormalizeTitle(unittest.TestCase):
    def test_collapses_whitespace(self):
        self.assertEqual(normalize_title("  hello   world  "), "hello world")


class TestParseDateFromText(unittest.TestCase):
    def test_month_name_format(self):
        text = "A report of an interview conducted on August 27, 2026, is attached."
        self.assertEqual(parse_date_from_text(text), "2026-08-27")

    def test_numeric_mm_dd_yyyy_format(self):
        text = "Deadline: 08-27-2026 for response."
        self.assertEqual(parse_date_from_text(text), "2026-08-27")

    def test_iso_format(self):
        text = "Filed on 2026-08-27 as instructed."
        self.assertEqual(parse_date_from_text(text), "2026-08-27")

    def test_no_date_returns_none(self):
        self.assertIsNone(parse_date_from_text("VIA E-MAIL ONLY\nDear Sirs,"))

    def test_invalid_calendar_date_is_skipped(self):
        # 13월은 존재하지 않으므로 무시하고, 그다음 유효한 날짜를 찾는다.
        text = "Ref. 13-40-2026 unrelated, actual date is August 27, 2026."
        self.assertEqual(parse_date_from_text(text), "2026-08-27")


class TestProcessSheetNewColumns(unittest.TestCase):
    def setUp(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "오전"

    def _set_row(self, row, *, subject="", management_no="", sender=""):
        self.ws.cell(row=row, column=COL_E).value = subject
        self.ws.cell(row=row, column=COL_I).value = management_no
        self.ws.cell(row=row, column=3).value = sender  # COL_C

    def test_blank_management_no_falls_back_to_gita_when_unclassified(self):
        self._set_row(2, subject="전혀 관련 없는 제목", management_no="")
        process_sheet(self.ws)
        self.assertEqual(self.ws.cell(row=2, column=COL_K).value, "기타")
        self.assertEqual(self.ws.cell(row=2, column=COL_L).value, "(기타)기타")

    def test_blank_management_no_does_not_override_existing_classification(self):
        self._set_row(2, subject="Automatic reply: out of office", management_no="")
        process_sheet(self.ws)
        self.assertEqual(self.ws.cell(row=2, column=COL_K).value, "ACK")

    def test_gita_fallback_skipped_for_empty_row(self):
        # 제목 없는 빈 행은 관리번호가 비어 있어도 손대지 않는다.
        self._set_row(2, subject="", management_no="")
        process_sheet(self.ws)
        self.assertIsNone(self.ws.cell(row=2, column=COL_K).value)

    def test_management_no_present_skips_gita_fallback(self):
        self._set_row(2, subject="전혀 관련 없는 제목", management_no="SP001")
        process_sheet(self.ws)
        self.assertIsNone(self.ws.cell(row=2, column=COL_K).value)

    def test_email_date_fills_m_column_on_subject_match(self):
        subject = "Report of Examiner Interview 08-27-2026"
        self._set_row(2, subject=subject, management_no="SP001")
        email_dates = {normalize_title(subject): "2026-08-27"}
        process_sheet(self.ws, email_dates)
        self.assertEqual(self.ws.cell(row=2, column=COL_M).value, "2026-08-27")

    def test_no_email_date_match_leaves_m_column_blank(self):
        self._set_row(2, subject="No matching email for this one", management_no="SP001")
        process_sheet(self.ws, email_dates={"다른 제목": "2026-08-27"})
        self.assertIsNone(self.ws.cell(row=2, column=COL_M).value)


if __name__ == "__main__":
    unittest.main()
