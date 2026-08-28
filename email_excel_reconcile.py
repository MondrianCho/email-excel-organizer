# -*- coding: utf-8 -*-
"""
오전 이메일 백업 폴더: 저장된 이메일 ↔ 엑셀 제목(헤더 '제목') 검사
- 이메일 파일 개수 vs 엑셀 2행~ 제목이 채워진 행 수 비교 → 차이 시 원인 분석
- 제목 집합 매칭(누락) + 엑셀 동일 제목 중복 행(두 번 접수) 보고
"""
from __future__ import annotations

import email
import email.policy
import sys
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("[오류] openpyxl이 없습니다. pip install openpyxl")
    sys.exit(1)

BACKUP_SUBDIR = "오전 이메일 백업"
HEADER_TITLE = "제목"
DATA_START_ROW = 2


def safe_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def normalize_title(s: str) -> str:
    """비교용: 앞뒤 공백 제거, 연속 공백 하나로."""
    if not s:
        return ""
    return " ".join(s.split())


def is_excel_file(p: Path) -> bool:
    if p.name.startswith("~$"):
        return False
    return p.suffix.lower() in (".xlsx", ".xlsm")


def find_title_column(ws) -> int | None:
    """1행에서 셀 값이 '제목'인 열 번호(1-based). 없으면 None."""
    max_c = min(ws.max_column or 1, 50)
    for col in range(1, max_c + 1):
        val = normalize_title(safe_str(ws.cell(row=1, column=col).value))
        if val == HEADER_TITLE:
            return col
    return None


def collect_excel_rows(folder: Path) -> tuple[list[tuple[str, str, str, int]], list[str]]:
    """
    엑셀 2행부터 제목이 비어 있지 않은 행마다 (정규화제목, 파일명, 시트명, 행번호) 수집.
    반환: rows, logs
    """
    rows: list[tuple[str, str, str, int]] = []
    logs: list[str] = []
    files = sorted(p for p in folder.iterdir() if p.is_file() and is_excel_file(p))
    if not files:
        logs.append("[안내] 엑셀 파일(.xlsx/.xlsm)이 없습니다.")
        return rows, logs

    for f in files:
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except Exception as e:
            logs.append(f"  [엑셀 열기 실패] {f.name}: {e}")
            continue
        try:
            sheet_rows = 0
            for ws in wb.worksheets:
                col = find_title_column(ws)
                if col is None:
                    logs.append(f"  [건너뜀] {f.name} / 시트 '{ws.title}': 1행에 '{HEADER_TITLE}' 헤더 없음")
                    continue
                for row in range(DATA_START_ROW, (ws.max_row or 0) + 1):
                    raw = ws.cell(row=row, column=col).value
                    t = normalize_title(safe_str(raw))
                    if t:
                        rows.append((t, f.name, ws.title, row))
                        sheet_rows += 1
            logs.append(f"  [엑셀] {f.name} — 제목 있는 행 {sheet_rows}건 수집")
        finally:
            wb.close()
    return rows, logs


def read_msg_subject(path: Path) -> str:
    try:
        import extract_msg
    except ImportError:
        raise RuntimeError("pip install extract-msg") from None
    msg = extract_msg.Message(str(path))
    try:
        return safe_str(msg.subject)
    finally:
        msg.close()


def read_eml_subject(path: Path) -> str:
    with path.open("rb") as f:
        p = email.message_from_binary_file(f, policy=email.policy.default)
    subj = p.get("Subject")
    if subj is None:
        return ""
    return safe_str(str(subj))


def collect_email_entries(
    folder: Path,
) -> tuple[list[tuple[str, str]], int, list[str], list[str]]:
    """
    각 이메일 파일당 (파일명, 정규화 제목) — 읽기 실패 시 제목은 빈 문자열.
    반환: entries, total_file_count(디스크상 .msg+.eml 개수), logs, errors
    """
    entries: list[tuple[str, str]] = []
    logs: list[str] = []
    errors: list[str] = []

    msg_files = sorted(folder.glob("*.msg"))
    eml_files = sorted(folder.glob("*.eml"))
    total_email_files = len(msg_files) + len(eml_files)

    has_extract = False
    try:
        import extract_msg  # noqa: F401

        has_extract = True
    except ImportError:
        pass

    msg_ok = 0
    if msg_files and not has_extract:
        errors.append(f"(전체 {len(msg_files)}개 .msg) pip install extract-msg 필요")
        for p in msg_files:
            entries.append((p.name, ""))
    else:
        for p in msg_files:
            try:
                s = normalize_title(read_msg_subject(p))
                entries.append((p.name, s))
                msg_ok += 1
            except Exception as e:
                errors.append(f"{p.name}: {e}")
                entries.append((p.name, ""))

    eml_ok = 0
    for p in eml_files:
        try:
            s = normalize_title(read_eml_subject(p))
            entries.append((p.name, s))
            eml_ok += 1
        except Exception as e:
            errors.append(f"{p.name}: {e}")
            entries.append((p.name, ""))

    if msg_files:
        logs.append(f"  [MSG] 파일 {len(msg_files)}개 (읽기 성공 {msg_ok}개)")
    if eml_files:
        logs.append(f"  [EML] 파일 {len(eml_files)}개 (읽기 성공 {eml_ok}개)")
    if not msg_files and not eml_files:
        logs.append("[안내] .msg 또는 .eml 파일이 없습니다.")

    return entries, total_email_files, logs, errors


def run(parent_folder: str) -> None:
    root = Path(parent_folder).expanduser().resolve()
    if not root.is_dir():
        print(f"[오류] 폴더가 없습니다: {root}")
        sys.exit(1)

    backup = root / BACKUP_SUBDIR
    if not backup.is_dir():
        print(f"[오류] 하위 폴더가 없습니다: {backup}")
        print(f"       상위 경로에 '{BACKUP_SUBDIR}' 폴더를 두었는지 확인하세요.")
        sys.exit(1)

    print(f"검사 폴더: {backup}\n")

    excel_rows, excel_logs = collect_excel_rows(backup)
    email_entries, email_file_count, email_logs, email_errors = collect_email_entries(backup)

    for line in excel_logs:
        print(line)
    print()
    for line in email_logs:
        print(line)
    if email_errors:
        print("\n[이메일 읽기 오류]")
        for e in email_errors:
            print(f"  - {e}")
    print()

    # 엑셀: 행 수, 고유 제목, 중복(동일 제목 여러 행)
    excel_data_rows = len(excel_rows)
    excel_title_counter = Counter(t for t, _, _, _ in excel_rows)
    excel_titles = set(excel_title_counter.keys())
    title_locations: dict[str, list[tuple[str, str, int]]] = defaultdict(list)
    for t, fn, sh, row in excel_rows:
        title_locations[t].append((fn, sh, row))

    # 이메일: 파일 수(디스크), 제목별 건수, 고유 제목
    email_subject_counter = Counter(s for _, s in email_entries if s)
    email_subjects = set(email_subject_counter.keys())
    empty_subject_files = [fn for fn, s in email_entries if not s]

    # 집합 차이
    only_email = sorted(email_subjects - excel_titles)
    only_excel = sorted(excel_titles - email_subjects)

    print("=" * 60)
    print("건수 비교 (이메일 파일 수 ↔ 엑셀 제목 열 실제 행 수, 2행부터)")
    print("=" * 60)
    print(f"이메일 파일(.msg+.eml): {email_file_count}개")
    print(f"엑셀 제목이 채워진 행: {excel_data_rows}건")
    print(f"엑셀 고유 제목 수: {len(excel_titles)}개 / 이메일 고유 제목 수(빈 제목 제외): {len(email_subjects)}개")
    print()

    dup_excel_titles = sorted((t, excel_title_counter[t]) for t in excel_titles if excel_title_counter[t] > 1)
    dup_email_subjects = sorted((s, email_subject_counter[s]) for s in email_subjects if email_subject_counter[s] > 1)

    if email_file_count != excel_data_rows:
        print(f"[차이] 파일 {email_file_count}개 vs 엑셀 행 {excel_data_rows}건 → {email_file_count - excel_data_rows:+d}")
        print("가능한 원인:")
        reasons: list[str] = []
        extra_excel_from_dup = sum(excel_title_counter[t] - 1 for t in excel_titles if excel_title_counter[t] > 1)
        if extra_excel_from_dup:
            reasons.append(f"엑셀에서 동일 제목이 여러 행이면 행 수만 늘어남 (중복 접수 {extra_excel_from_dup}행)")
        same_subj_extra_files = sum(email_subject_counter[s] - 1 for s in email_subjects if email_subject_counter[s] > 1)
        if same_subj_extra_files:
            reasons.append(f"제목이 같은 이메일 파일이 여러 개면 파일 수만 늘어남 (+{same_subj_extra_files}개)")
        if only_email:
            reasons.append(f"엑셀에 없는 제목의 이메일 {len(only_email)}개 → 엑셀 행이 적을 수 있음")
        if only_excel:
            reasons.append(f"이메일과 안 맞는 엑셀 제목 {len(only_excel)}개 → 엑셀 행만 많을 수 있음")
        if empty_subject_files:
            reasons.append(f"제목을 못 읽었거나 빈 제목 이메일 {len(empty_subject_files)}개 (건수·매칭에 불리)")
        if email_errors:
            reasons.append("일부 .msg/.eml 읽기 실패 시 실제 제목 반영이 어긋날 수 있음")
        if not reasons:
            reasons.append("위 항목 외 조합(여러 원인 동시)일 수 있음 — 아래 누락·중복 목록을 함께 확인하세요.")
        for r in reasons:
            print(f"  · {r}")
        print()
    else:
        print("[OK] 이메일 파일 수와 엑셀 제목 행 수가 같습니다.\n")

    if dup_excel_titles:
        print(f"[중복 접수] 엑셀에 동일 제목이 2행 이상 ({len(dup_excel_titles)}종류)")
        for title, cnt in dup_excel_titles:
            print(f"  · 「{title}」 → {cnt}행")
            for fn, sh, row in title_locations[title]:
                print(f"      - {fn} / 시트「{sh}」 {row}행")
        print()
    else:
        print("[OK] 엑셀에서 동일 제목 중복 행 없음\n")

    if dup_email_subjects:
        print(f"[참고] 동일 제목의 이메일 파일이 여러 개 ({len(dup_email_subjects)}종류)")
        for subj, cnt in dup_email_subjects:
            names = [fn for fn, s in email_entries if s == subj]
            print(f"  · 「{subj}」 → {cnt}개 파일: {', '.join(names)}")
        print()

    print("=" * 60)
    print("제목 매칭 (정규화 후 완전 일치)")
    print("=" * 60)

    if only_email:
        print(f"[누락] 엑셀에 없는 이메일 제목 ({len(only_email)}개)")
        for t in only_email:
            print(f"  - {t}")
        print()
    else:
        print("[OK] 이메일에만 있고 엑셀에 없는 제목 없음\n")

    if only_excel:
        print(f"[누락] 이메일과 매칭되지 않는 엑셀 제목 ({len(only_excel)}개)")
        for t in only_excel:
            print(f"  - {t}")
        print()
    else:
        print("[OK] 엑셀에만 있고 이메일 제목과 다른 항목 없음\n")

    if empty_subject_files:
        print(f"[참고] 제목이 비어 있거나 읽지 못한 이메일 파일 ({len(empty_subject_files)}개) — 수동 확인 권장")
        for fn in empty_subject_files:
            print(f"  - {fn}")
        print()

    if (
        not only_email
        and not only_excel
        and not dup_excel_titles
        and email_file_count == excel_data_rows
    ):
        print("건수·집합·엑셀 중복 모두 일치합니다.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python email_excel_reconcile.py \"상위폴더경로\"")
        print(f"       → 해당 폴더 안의 '{BACKUP_SUBDIR}' 를 검사합니다.")
        print('예: python email_excel_reconcile.py "C:\\\\Users\\\\USER\\\\Desktop\\\\접수"')
        sys.exit(1)
    run(sys.argv[1])
