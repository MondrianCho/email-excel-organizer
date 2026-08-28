# -*- coding: utf-8 -*-
"""
이메일 접수 엑셀 정리 자동화
- 폴더 내 엑셀 파일명을 'yyyy-mm-dd 이메일 접수_초안.xlsx'로 변경
- E열(제목) 2행부터 확인하여 K열(업무종류), L열, N열, O열 자동 입력
- '오전 이메일 백업' 폴더의 이메일 본문 앞부분에서 날짜를 찾아 M열(마감일)에 입력
- I열(관리번호)이 빈칸이고 K열이 미정이면 K열 '기타' / L열 '(기타)기타'로 보완
"""
import email
import email.policy
import html
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
except ImportError:
    print("[오류] openpyxl이 설치되어 있지 않습니다. 다음 명령으로 설치하세요:")
    print("  pip install openpyxl")
    sys.exit(1)

# 열 번호 (1-based)
COL_C = 3   # 발신
COL_E = 5   # 제목
COL_I = 9   # 관리번호
COL_K = 11  # 업무종류
COL_L = 12
COL_M = 13  # 마감일
COL_N = 14
COL_O = 15

DATA_START_ROW = 2

BACKUP_SUBDIR = "오전 이메일 백업"
BODY_SCAN_LINES = 5

# 이메일 접수 구조(제목/관리번호/업무종류 등)와 일치하는 시트만 자동 분류 대상으로 삼는다.
# 그 외 시트(사건리스트, 수식 등)는 건드리지 않는다.
TARGET_SHEET_NAMES = {"오전", "오후"}

MONTH_NAMES = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}

# "August 27, 2026" / "Aug. 27 2026"
RE_MONTH_DD_YYYY = re.compile(r"\b([A-Za-z]{3,9})\.?\s+(\d{1,2}),?\s+(\d{4})\b")
# "27 August 2026"
RE_DD_MONTH_YYYY = re.compile(r"\b(\d{1,2})\s+([A-Za-z]{3,9})\.?,?\s+(\d{4})\b")
# "08-27-2026" / "08/27/2026" (월/일/년, 미국식)
RE_MM_DD_YYYY = re.compile(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{4})\b")
# "2026-08-27"
RE_YYYY_MM_DD = re.compile(r"\b(\d{4})-(\d{1,2})-(\d{1,2})\b")

# 시트 글꼴: 맑은 고딕 10
SHEET_FONT = Font(name="맑은 고딕", size=10)

# (E열 검사 문자열, 대소문자 무시), (K값, L값, N값, O값)
# 더 긴/구체적인 문구를 먼저 두어 우선 매칭
RULES_E_TITLE = [
    # ACK
    ("ACK", "ACK", None, None, None),
    ("Automatic reply", "ACK", None, None, None),
    ("AutoResponse", "ACK", None, None, None),
    # OA (L,N,O 있는 것 먼저)
    ("Form 3 deadline", "OA", "(OA)마감RMD", None, "FORM3 RMD"),
    ("information on the status of the file", "OA", None, None, "상태문의답변통지서"),
    ("New case for foreign filing", "OA", "(OA)해외연구소", None, None),
    ("Request for Non-Provisional Appln.", "OA", "(OA)해외연구소", None, None),
    ("Request for CS Appln.", "OA", "(OA)해외연구소", None, None),
    # OA (L값 있는 것 먼저)
    ("COMMENT", "OA", "(OA)의견(analysis)", None, None),
    ("Draft Response", "OA", "(OA)의견(analysis)", None, None),
    ("Recommendations", "OA", "(OA)의견(analysis)", None, None),
    ("Analysis and Proposed Amendment", "OA", "(OA)의견(analysis)", None, None),
    ("Interview Schedule", "OA", "(OA)인터뷰", None, None),
    ("INTERVIEW", "OA", "(OA)인터뷰", None, None),
    # OA (일반) / 비용 예외: Maintenance Reminder는 reminder보다 먼저 배치
    ("Maintenance Reminder", "비용", "(비용)연차료/RMD", None, None),
    ("reminder", "OA", None, None, None),
    ("Received Non-Final Office Action", "OA", None, None, None),
    ("Non-Final Office Action", "OA", None, None, None),
    ("forwarding Office Action", "OA", None, None, None),
    ("forwarding Final Office Action", "OA", None, None, None),
    ("Reporting Summons to attend Oral Proceedings", "OA", None, None, None),
    ("Report search report", "OA", None, None, None),
    ("First Examination Report", "OA", None, None, None),
    ("Report office action", "OA", None, None, None),
    ("EESR", "OA", None, None, None),
    # 등록
    ("DECISION TO GRANT", "등록", "(등록)EP_decision to grant", "비용팀확인", None),
    ("Report grant", "등록", "(등록)EP_decision to grant", "비용팀확인", None),
    ("Rule 71(3) EPC received", "등록", None, None, None),
    ("report R. 71-3", "등록", None, None, None),
    ("ISSUE NOTIFICATION", "등록", "(등록)US_issue notification", None, None),
    ("Reporting Notice of Allowance", "등록", None, None, None),
    ("Report Corrected Notice of Allowability", "등록", None, None, "Corrected NOA"),
    ("Received Corrected Notice of Allowance", "등록", None, None, None),
    ("Notice of Allowance", "등록", None, None, None),
    ("report Rule 71(3)", "등록", None, None, None),
    ("report intention to grant", "등록", None, None, None),
    # 비용 (L,N 있는 것 먼저)
    ("Entering into Sub.Exam", "비용", "(비용)CN_실질심사착수통지", "OA팀확인", None),
    ("Entering the Substantive Examination Stage", "비용", "(비용)CN_실질심사착수통지", "OA팀확인", None),
    # IDS (INVOICE보다 먼저 배치: 제목에 IDS 포함 시 IDS 우선 분류)
    ("Information Disclosure Statement", "IDS", "(IDS)IDS제출완료", None, None),
    ("Filed IDS", "IDS", "(IDS)IDS제출완료", None, None),
    ("IDS FILED", "IDS", "(IDS)IDS제출완료", None, None),
    ("FILED – IDS", "IDS", "(IDS)IDS제출완료", None, None),
    ("FILED - IDS", "IDS", "(IDS)IDS제출완료", None, None),
    # 비용 (일반)
    ("INVOICE", "비용", None, None, None),
    ("certificate", "비용", None, None, None),
    ("Received Official Filing Receipt", "비용", None, None, None),
    ("Received OFR", "비용", None, None, None),
    ("report no opposition", "비용", "(비용)EP_이의신청기간만료통지", None, None),
    ("Issue fee paid", "비용", None, None, None),
    ("opposition notification", "비용", None, None, None),
    ("abandonment", "비용", None, None, None),
    ("Filed RCE", "비용", None, None, None),
    ("Filed Amendment", "비용", None, None, None),
    ("filing report", "비용", None, None, None),
    ("Report Filing", "비용", None, None, None),
    ("Annuity", "비용", None, None, None),
    ("FORM3 FILED", "비용", None, None, None),
    # 출원
    ("ASSIGNMENT WITH DECLARATION", "출원", None, None, None),
    ("번역문", "출원", None, None, None),
    # 기타
    ("상용파트 사건현황리스트", "기타", "(기타)당소발송건", None, None),
    ("IP출원3그룹 사건현황리스트", "기타", "(기타)당소발송건", None, None),
    ("Newsletter", "기타", None, None, None),
]

# C열 발신 규칙: (발신 문자열 포함 시), K값, L값
RULES_C_SENDER = [
    ("담당자A", "OA", "(OA)해외연구소"),
    ("epct-noreply@wipo.int", "PCT서류", None),
    ("담당자B", "출원", "(출원)외주번역/번역문청구서"),
]


def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()


def normalize_title(s: str) -> str:
    """비교용: 앞뒤 공백 제거, 연속 공백 하나로."""
    if not s:
        return ""
    return " ".join(s.split())


def get_e_value(ws, row):
    return safe_str(ws.cell(row=row, column=COL_E).value)


def get_c_value(ws, row):
    return safe_str(ws.cell(row=row, column=COL_C).value)


def get_i_value(ws, row):
    return safe_str(ws.cell(row=row, column=COL_I).value)


def match_e_rules(title):
    """E열 제목으로 K,L,N,O 값 결정. (K, L, N, O) 또는 None."""
    title_upper = title.upper()
    title_lower = title.lower()
    for keyword, k_val, l_val, n_val, o_val in RULES_E_TITLE:
        # 대소문자 구분 없이 포함 여부 (한글은 그대로)
        if keyword in title or keyword.upper() in title_upper or keyword.lower() in title_lower:
            return (k_val, l_val, n_val, o_val)
    return None


def _parse_month(name: str):
    return MONTH_NAMES.get(name.lower().rstrip("."))


def parse_date_from_text(text: str):
    """텍스트에서 처음 발견되는 유효한 날짜를 yyyy-mm-dd 문자열로 반환. 없으면 None."""
    candidates = []
    for m in RE_MONTH_DD_YYYY.finditer(text):
        month = _parse_month(m.group(1))
        if month is not None:
            candidates.append((m.start(), month, int(m.group(2)), int(m.group(3))))
    for m in RE_DD_MONTH_YYYY.finditer(text):
        month = _parse_month(m.group(2))
        if month is not None:
            candidates.append((m.start(), month, int(m.group(1)), int(m.group(3))))
    for m in RE_YYYY_MM_DD.finditer(text):
        candidates.append((m.start(), int(m.group(2)), int(m.group(3)), int(m.group(1))))
    for m in RE_MM_DD_YYYY.finditer(text):
        candidates.append((m.start(), int(m.group(1)), int(m.group(2)), int(m.group(3))))

    candidates.sort(key=lambda c: c[0])
    for _, month, day, year in candidates:
        try:
            return date(year, month, day).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def is_email_file(p: Path) -> bool:
    return p.suffix.lower() in (".eml", ".msg")


def strip_html(text: str) -> str:
    text = re.sub(r"(?is)<(script|style)[^>]*>.*?</\1>", " ", text)
    text = re.sub(r"(?s)<[^>]+>", " ", text)
    return html.unescape(text)


def read_eml_subject_and_lines(path: Path, n: int = BODY_SCAN_LINES):
    with path.open("rb") as fh:
        msg = email.message_from_binary_file(fh, policy=email.policy.default)
    subject = normalize_title(safe_str(msg.get("Subject")))

    part = msg.get_body(preferencelist=("plain", "html"))
    text = ""
    if part is not None:
        try:
            text = part.get_content()
        except Exception:
            text = ""
        if part.get_content_type() == "text/html":
            text = strip_html(text)
    lines = [l.strip() for l in text.splitlines() if l.strip()][:n]
    return subject, lines


def read_msg_subject_and_lines(path: Path, n: int = BODY_SCAN_LINES):
    import extract_msg

    msg = extract_msg.Message(str(path))
    try:
        subject = normalize_title(safe_str(msg.subject))
        text = safe_str(msg.body)
        if not text:
            html_body = msg.htmlBody
            if isinstance(html_body, bytes):
                html_body = html_body.decode("utf-8", errors="ignore")
            text = strip_html(html_body) if html_body else ""
    finally:
        msg.close()
    lines = [l.strip() for l in text.splitlines() if l.strip()][:n]
    return subject, lines


def collect_email_subject_dates(backup_folder: Path):
    """'오전 이메일 백업' 폴더의 이메일을 읽어 (정규화 제목 → 본문 앞부분 날짜) 매핑 생성.
    같은 제목의 이메일이 여러 개면 어느 행에 대응하는지 알 수 없으므로 제외한다."""
    logs = []

    if not backup_folder.is_dir():
        logs.append(f"[안내] '{BACKUP_SUBDIR}' 폴더가 없어 M열(마감일) 자동 입력은 건너뜁니다.")
        return {}, logs

    msg_files = sorted(backup_folder.glob("*.msg"))
    eml_files = sorted(backup_folder.glob("*.eml"))

    has_extract = False
    try:
        import extract_msg  # noqa: F401

        has_extract = True
    except ImportError:
        pass
    if msg_files and not has_extract:
        logs.append(f"[안내] .msg 파일 {len(msg_files)}개 — pip install extract-msg 필요, 건너뜁니다.")
        msg_files = []

    subject_date: dict = {}
    subject_count: dict = defaultdict(int)

    for p in msg_files + eml_files:
        try:
            if p.suffix.lower() == ".msg":
                subject, lines = read_msg_subject_and_lines(p)
            else:
                subject, lines = read_eml_subject_and_lines(p)
        except Exception as e:
            logs.append(f"  [오류] {p.name}: {e}")
            continue

        if not subject:
            continue

        found_date = parse_date_from_text("\n".join(lines))
        subject_count[subject] += 1
        if subject not in subject_date:
            subject_date[subject] = found_date

    duplicates = [s for s, c in subject_count.items() if c > 1]
    if duplicates:
        logs.append(f"[안내] 동일 제목 이메일 {len(duplicates)}종류는 행을 특정할 수 없어 M열 입력에서 제외합니다.")

    result = {
        s: d for s, d in subject_date.items()
        if d is not None and subject_count[s] == 1
    }
    logs.append(f"[안내] 이메일 {len(msg_files) + len(eml_files)}개 중 날짜 매칭 {len(result)}건 확보.")
    return result, logs


def process_sheet(ws, email_dates=None):
    """시트에서 E열 2행부터 읽어 K,L,M,N,O 채우기."""
    email_dates = email_dates or {}
    max_row = ws.max_row
    if max_row < DATA_START_ROW:
        return 0

    changed = 0
    for row in range(DATA_START_ROW, max_row + 1):
        c_val = get_c_value(ws, row)
        e_val = get_e_value(ws, row)

        k_val, l_val, n_val, o_val = None, None, None, None

        # 1) C열 발신 규칙 (담당자A → OA + (OA)해외연구소, epct-noreply@wipo.int → PCT서류 등)
        if c_val:
            for sender_key, ck, cl in RULES_C_SENDER:
                if sender_key in c_val:
                    k_val = ck
                    if cl is not None:
                        l_val = cl
                    break

        # 2) E열 제목 규칙 (C 규칙으로 이미 정해진 경우에도 E 규칙이 있으면 L,N,O 추가 적용)
        #    표 기준으로는 C조건이 별도 행이므로, C조건이면 그대로 두고 E는 보조로 쓸 수 있음.
        #    여기서는 C가 담당자A이면 이미 K,L 설정했으므로 E는 추가 L,N,O만 채우거나 유지.
        e_matched = match_e_rules(e_val) if e_val else None
        if e_matched:
            ek, el, en, eo = e_matched
            if k_val is None:
                k_val = ek
            if el is not None:
                l_val = el
            if en is not None:
                n_val = en
            if eo is not None:
                o_val = eo

        # 3) 실제 접수 행(제목 있음)인데 관리번호(I열)가 빈칸이고
        #    위 규칙으로 K열이 정해지지 않았으면 '기타'로 보완
        if e_val and k_val is None and not get_i_value(ws, row):
            k_val = "기타"
            l_val = "(기타)기타"

        # 4) 제목이 이메일 본문에서 찾은 날짜와 매칭되면 M열(마감일)에 반영
        m_val = email_dates.get(normalize_title(e_val)) if e_val else None

        # 셀에 반영 (값이 정해진 것만)
        if k_val is not None:
            if ws.cell(row=row, column=COL_K).value != k_val:
                ws.cell(row=row, column=COL_K).value = k_val
                changed += 1
        if l_val is not None:
            if ws.cell(row=row, column=COL_L).value != l_val:
                ws.cell(row=row, column=COL_L).value = l_val
                changed += 1
        if m_val is not None:
            if ws.cell(row=row, column=COL_M).value != m_val:
                ws.cell(row=row, column=COL_M).value = m_val
                changed += 1
        if n_val is not None:
            if ws.cell(row=row, column=COL_N).value != n_val:
                ws.cell(row=row, column=COL_N).value = n_val
                changed += 1
        if o_val is not None:
            if ws.cell(row=row, column=COL_O).value != o_val:
                ws.cell(row=row, column=COL_O).value = o_val
                changed += 1

    return changed


def apply_sheet_font(ws):
    """시트 사용 영역 전체에 맑은 고딕 10pt 적용."""
    if ws.max_row < 1 or ws.max_column < 1:
        return
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = SHEET_FONT


def is_excel_file(p: Path) -> bool:
    if p.name.startswith("~$"):
        return False
    return p.suffix.lower() in (".xlsx", ".xlsm")


def get_new_filename(file_path: Path, use_mod_date: bool = True) -> str:
    """작성일 yyyy-mm-dd 이메일 접수_초안.xlsx 형식 이름 반환."""
    if use_mod_date:
        mtime = file_path.stat().st_mtime
        dt = datetime.fromtimestamp(mtime)
    else:
        dt = datetime.now()
    date_str = dt.strftime("%Y-%m-%d")
    return f"{date_str} 이메일 접수_초안.xlsx"


def process_folder(folder_path: str):
    folder = Path(folder_path).expanduser().resolve()
    if not folder.exists() or not folder.is_dir():
        print(f"[오류] 폴더가 존재하지 않습니다: {folder}")
        sys.exit(1)

    files = [p for p in folder.iterdir() if p.is_file() and is_excel_file(p)]
    if not files:
        print(f"[안내] 해당 폴더에서 엑셀 파일을 찾지 못했습니다: {folder}")
        return

    email_dates, email_logs = collect_email_subject_dates(folder / BACKUP_SUBDIR)
    for line in email_logs:
        print(line)

    print(f"\n[안내] 엑셀 파일 {len(files)}개 발견. 처리 시작.\n")

    for f in files:
        try:
            wb = openpyxl.load_workbook(f, read_only=False, data_only=False)
            total_changed = 0
            for ws in wb.worksheets:
                if ws.title not in TARGET_SHEET_NAMES:
                    continue
                total_changed += process_sheet(ws, email_dates)
                apply_sheet_font(ws)
            wb.save(f)
            wb.close()
            print(f"  [OK] {f.name} — {total_changed}개 셀 반영")
        except Exception as e:
            print(f"  [오류] {f.name} — {e}")
            continue

        # 파일명 변경: 작성일 yyyy-mm-dd 이메일 접수_초안.xlsx
        new_name = get_new_filename(f, use_mod_date=True)
        new_path = f.parent / new_name
        if new_path == f:
            continue
        if new_path.exists() and new_path != f:
            # 이미 같은 이름 파일 있으면 번호 붙이기
            for i in range(1, 100):
                candidate = f.parent / f"{new_name.replace('.xlsx','')}_{i}.xlsx"
                if not candidate.exists():
                    new_path = candidate
                    break
        try:
            f.rename(new_path)
            print(f"       → 파일명 변경: {new_path.name}")
        except Exception as e:
            print(f"       → 파일명 변경 실패: {e}")

    print("\n[완료] 정리 작업이 끝났습니다.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python email_excel_organize.py \"폴더경로\"")
        print("예: python email_excel_organize.py \"C:\\Users\\Erin\\Desktop\\이메일접수\"")
        sys.exit(1)
    process_folder(sys.argv[1])
